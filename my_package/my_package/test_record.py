import os
from openpyxl import load_workbook, Workbook
import numpy as np
from sklearn.metrics import roc_auc_score, roc_curve, f1_score
import torch
import torch.nn.functional as F
from typing import Union
from PIL import Image
import seaborn as sns
from matplotlib import pyplot as plt

class TestRecord():
    def __init__(self,
                 model_name:str,
                 label_names:Union[list[str], None],
                 output_path = "/home/chengfangchi/graduate2302/code/cfc/最后的冯如杯"):
        excel_path = os.path.join(output_path, "record.xlsx")
        if not os.path.exists(excel_path):
            work_book = Workbook()
            work_book.save(excel_path)

        self.excel_path = excel_path
        self.img_path = os.path.join(output_path, "record.png")
        self.model_name = model_name
        self.has_end = False
        self.preds = []
        self.results = []
        self.labels = []
        self.label_names = label_names

    def record(self, logits:torch.Tensor, label:int):
        preds = F.softmax(logits, -1)
        self.preds.append(preds)
        self.labels.append(label)
        self.results.append(preds.argmax(dim=-1).item())

    def end_record(self):
        assert not self.has_end
        self.has_end = True

        preds = torch.stack(self.preds)
        results = torch.tensor(self.results)
        labels = torch.tensor(self.labels)

        if self.label_names is None: self.label_names = [str(i) for i in range(preds.shape[1])]
        work_book = load_workbook(self.excel_path)
        if "main" not in work_book.sheetnames:
            work_book.create_sheet("main")
            work_sheet = work_book["main"]

            work_sheet.append(["model_name", "acc", "micro_auroc", "macro_auroc", "micro_f1", "macro_f1"]+self.label_names)
            work_book.save(self.excel_path) # type: ignore

        total_num = results.shape[0]
        correct_num = torch.sum(results == labels).item() # type: ignore
        acc = round(correct_num / total_num * 100, 1)

        classes_acc = []
        for i in range(preds.shape[1]):
            class_correct = torch.sum(results[labels==i]==labels[labels==i]).item()
            class_total = torch.sum(labels==i).item()
            classes_acc.append(round(class_correct/class_total*100,1))

        self.get_confusion_matrix_img()

        auroc_dict = self.get_auroc(np.array(preds), np.array(labels), self.label_names)
        micro_auroc, macro_auroc = auroc_dict["Micro-average"], auroc_dict["Macro-average"]

        micro_f1 = f1_score(self.labels, self.results, average="micro")
        macro_f1 = f1_score(self.labels, self.results, average="macro")

        work_book = load_workbook(self.excel_path)
        work_sheet = work_book["main"]
        work_sheet.append([self.model_name, acc, micro_auroc, macro_auroc, micro_f1, macro_f1] + classes_acc)
        work_book.save(self.excel_path)

    @staticmethod
    def get_auroc(preds:np.ndarray, labels:np.ndarray, class_name:list[str])->dict:
        data_num, class_num = preds.shape
        assert len(class_name) == class_num

        one_hot_labels = np.zeros_like(preds, dtype=np.int64)
        one_hot_labels[np.arange(data_num), labels] = 1

        # 计算每个类别的 ROC 曲线和 AUC 值
        fpr = dict()
        tpr = dict()
        roc_auc = dict()
        for i in range(class_num):
            fpr[i], tpr[i], _ = roc_curve(one_hot_labels[:, i], preds[:, i]) # type: ignore
            roc_auc[i] = roc_auc_score(one_hot_labels[:, i], preds[:, i]) # type: ignore

        # 计算微观平均 ROC 曲线和 AUC 值
        roc_auc_micro = roc_auc_score(one_hot_labels, preds, average="micro")

        # 计算宏观平均 ROC 曲线和 AUC 值
        all_fpr = np.unique(np.concatenate([fpr[i] for i in range(class_num)]))
        mean_tpr = np.zeros_like(all_fpr)
        for i in range(class_num):
            mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])
        mean_tpr /= class_num
        roc_auc_macro = roc_auc_score(one_hot_labels, preds, average="macro")

        auroc_dict = {"Micro-average":float(roc_auc_micro), "Macro-average":float(roc_auc_macro)}
        for index, name in enumerate(class_name):
            auroc_dict[name] = float(roc_auc[index])
        return auroc_dict

    def get_confusion_matrix_img(self):
        confusion_matrix = torch.zeros((self.preds[0].shape[0], self.preds[0].shape[0]))
        for i, j in zip(self.results, self.labels):
            confusion_matrix[j, i] += 1
        confusion_matrix = confusion_matrix / torch.sum(confusion_matrix, dim=-1, keepdim=True)

        height, width = confusion_matrix.shape
        plt.figure(figsize=(width, height))
        sns.heatmap(confusion_matrix, annot=True, fmt=".2f", cmap='viridis', vmin=0, vmax=1)
        plt.title(self.model_name)
        plt.xlabel("predict")
        plt.ylabel("true")

        if os.path.exists(self.img_path):
            origin_img = Image.open(self.img_path)
            height, width = origin_img.height, origin_img.width
            origin_img = np.array(origin_img)
            plt.savefig(self.img_path)
            new_img = Image.open(self.img_path)
            new_width = width + new_img.width
            new_height = max(height, new_img.height)
            merged_img = Image.new('RGB', (new_width, new_height))
            merged_img.paste(Image.fromarray(origin_img), (0, 0))
            merged_img.paste(new_img, (width, 0))
            merged_img.save(self.img_path)
        else:
            plt.savefig(self.img_path)

if __name__ == "__main__":
    import random
    test_record = TestRecord("test_model", ["0", "1", "2", "3", "4"])
    for i in range(1000):
        test_record.record(torch.randn((5, )), random.randint(0, 4))
    test_record.end_record()